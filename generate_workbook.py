"""
Armageddon Treats — Comprehensive Financial Projections Workbook
Generates a multi-sheet Excel workbook with realistic projections grounded
in the business plan audit data (SBA 7(a) $14k microloan, mobile dessert
vendor at Santa Monica Farmers Markets).
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from copy import copy
import math

# ══════════════════════════════════════════════════════════════
#  BRAND PALETTE & STYLES
# ══════════════════════════════════════════════════════════════
BLACK       = "1A1A1A"
DARK_GRAY   = "2D2D2D"
MID_GRAY    = "4A4A4A"
LIGHT_GRAY  = "F2F2F2"
ACCENT_RED  = "C0392B"
ACCENT_DARK = "962D22"
WHITE       = "FFFFFF"
PALE_RED    = "FADBD8"
PALE_GREEN  = "D5F5E3"
PALE_YELLOW = "FEF9E7"
PALE_BLUE   = "D6EAF8"
MEDIUM_BLUE = "2980B9"
DARK_GREEN  = "1E8449"
WARM_GRAY   = "F8F8F8"

FONT_NAME = "Calibri"

# Reusable style objects
font_title     = Font(name=FONT_NAME, size=16, bold=True, color=BLACK)
font_subtitle  = Font(name=FONT_NAME, size=12, bold=True, color=MID_GRAY)
font_section   = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_RED)
font_header    = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
font_body      = Font(name=FONT_NAME, size=10, color=DARK_GRAY)
font_body_bold = Font(name=FONT_NAME, size=10, bold=True, color=DARK_GRAY)
font_small     = Font(name=FONT_NAME, size=9, color=MID_GRAY)
font_small_italic = Font(name=FONT_NAME, size=9, italic=True, color=MID_GRAY)
font_currency  = Font(name=FONT_NAME, size=10, color=DARK_GRAY)
font_negative  = Font(name=FONT_NAME, size=10, color=ACCENT_RED)
font_positive  = Font(name=FONT_NAME, size=10, color=DARK_GREEN)
font_total     = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
font_label_sm  = Font(name=FONT_NAME, size=9, bold=True, color=MID_GRAY)

fill_header    = PatternFill("solid", fgColor=BLACK)
fill_subheader = PatternFill("solid", fgColor=DARK_GRAY)
fill_accent    = PatternFill("solid", fgColor=ACCENT_RED)
fill_light     = PatternFill("solid", fgColor=LIGHT_GRAY)
fill_white     = PatternFill("solid", fgColor=WHITE)
fill_pale_red  = PatternFill("solid", fgColor=PALE_RED)
fill_pale_green= PatternFill("solid", fgColor=PALE_GREEN)
fill_pale_yellow=PatternFill("solid", fgColor=PALE_YELLOW)
fill_pale_blue = PatternFill("solid", fgColor=PALE_BLUE)
fill_warm      = PatternFill("solid", fgColor=WARM_GRAY)
fill_total     = PatternFill("solid", fgColor=ACCENT_DARK)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
align_right  = Alignment(horizontal="right", vertical="center")
align_wrap   = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

thick_bottom = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="medium", color=BLACK),
)

FMT_USD      = '"$"#,##0'
FMT_USD_DEC  = '"$"#,##0.00'
FMT_PCT      = '0.0%'
FMT_PCT_1    = '0%'
FMT_NUM      = '#,##0'
FMT_NUM_DEC  = '#,##0.0'


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════
def style_header_row(ws, row, cols, fill=None, font=None):
    """Style a header row across specified columns."""
    f = fill or fill_header
    fn = font or font_header
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = f
        cell.font = fn
        cell.alignment = align_center
        cell.border = thin_border


def style_data_cell(cell, fmt=None, bold=False, fill=None):
    """Style a single data cell."""
    cell.font = font_body_bold if bold else font_body
    cell.border = thin_border
    cell.alignment = align_right if fmt else align_left
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill


def write_row(ws, row, data, start_col=1, fmt=None, bold=False, fill=None):
    """Write a list of values to a row with formatting."""
    for i, val in enumerate(data):
        # Skip None values — leave cell empty without writing a string
        if val is None:
            cell = ws.cell(row=row, column=start_col + i)
            cell.border = thin_border
            fl = fill[i] if isinstance(fill, list) else fill
            if fl:
                cell.fill = fl
            continue
        cell = ws.cell(row=row, column=start_col + i, value=val)
        f = fmt[i] if isinstance(fmt, list) else fmt
        b = bold[i] if isinstance(bold, list) else bold
        fl = fill[i] if isinstance(fill, list) else fill
        style_data_cell(cell, fmt=f, bold=b, fill=fl)


def write_section_label(ws, row, col, text, merge_end=None):
    """Write a red section label."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font_section
    cell.alignment = align_left
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)


def write_title(ws, row, col, text, merge_end=None):
    """Write the sheet title."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font_title
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)


def write_subtitle(ws, row, col, text, merge_end=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font_subtitle
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)


def write_note(ws, row, col, text, merge_end=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font_small_italic
    cell.alignment = align_wrap
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end)


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def style_total_row(ws, row, cols, fill=None):
    f = fill or fill_total
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = f
        cell.font = font_total
        cell.border = thick_bottom
        cell.alignment = align_right


# ══════════════════════════════════════════════════════════════
#  CORE ASSUMPTIONS (grounded in the audit)
# ══════════════════════════════════════════════════════════════

# --- Startup / Capital ---
SBA_LOAN = 14000
INTEREST_RATE = 0.1325  # max 13.25%
LOAN_TERM_MONTHS = 120  # 10 years

# --- Products ---
COTTON_CANDY_PRICE = 5.50
COTTON_CANDY_COGS  = 0.35
SHAVED_ICE_PRICE   = 5.00
SHAVED_ICE_COGS    = 0.60
PRODUCT_MIX_CC     = 0.40   # 40% cotton candy, 60% shaved ice

# --- Operations ---
# Year 1: ramp-up. RFP in March, maybe approved by May/June.
# Months 1-3 (Jan-Mar): prep, permits, no market revenue - maybe a few private gigs
# Months 4-5: potential pop-ups, swap meets while awaiting RFP result
# Months 6+: market access (if approved), ramp up

# Days per week at market
MARKET_DAYS_Y1_RAMP = 1      # first months at market
MARKET_DAYS_Y1_FULL = 2      # later Y1
MARKET_DAYS_Y2      = 2.5    # adding a Wednesday or pop-up
MARKET_DAYS_Y3      = 3      # three markets/week

# Units sold per market day (realistic, not fantasy)
UNITS_PER_DAY_LOW    = 40    # worst case / slow day
UNITS_PER_DAY_BASE   = 70    # base case average
UNITS_PER_DAY_HIGH   = 120   # peak summer Saturday

# Seasonal multipliers (Jan=1 index, shaved ice is very seasonal in SoCal)
SEASONALITY = {
    1: 0.50, 2: 0.55, 3: 0.65, 4: 0.80, 5: 0.90, 6: 1.00,
    7: 1.15, 8: 1.15, 9: 1.00, 10: 0.80, 11: 0.60, 12: 0.50
}

# Stall fee (% of gross)
STALL_FEE_PCT = 0.125  # 12.5% average

# --- Fixed Monthly OPEX ---
COMMISSARY_RENT     = 500    # parking + wash-down tier
INSURANCE_GL        = 120
INSURANCE_AUTO      = 200
FUEL_BASE           = 350
GENERATOR_FUEL      = 75
MISC_MAINTENANCE    = 125
PHONE_PAYMENT_PROC  = 50    # Square/POS + phone
PACKAGING_OVERHEAD  = 40    # beyond per-unit COGS (napkins, bags, etc.)

# --- Startup Costs (from audit) ---
STARTUP_COSTS = {
    "Concession Trailer (10ft import)":     7500,
    "Equipment & Branding (wrap, machines)": 2000,
    "Permits & Health Licensing":           1200,
    "Initial Insurance Premiums":           1000,
    "Initial Inventory & Packaging":         800,
    "Commissary Deposit / First Month":      500,
    "Working Capital / Contingency":        1000,
}

# --- Private Events / Side Revenue ---
PRIVATE_EVENT_REVENUE = 400   # net per event
PRIVATE_EVENTS_Y1 = [0,0,0,0,1,1,1,2,2,1,1,2]  # per month
PRIVATE_EVENTS_Y2 = [1,1,1,2,2,3,3,4,3,2,2,3]
PRIVATE_EVENTS_Y3 = [2,2,2,3,3,4,4,5,4,3,3,4]


# ══════════════════════════════════════════════════════════════
#  COMPUTATION ENGINE
# ══════════════════════════════════════════════════════════════

def calc_monthly_payment(principal, annual_rate, months):
    """Standard amortization formula."""
    r = annual_rate / 12
    if r == 0:
        return principal / months
    return principal * (r * (1 + r)**months) / ((1 + r)**months - 1)


def calc_blended_price():
    """Weighted average price and COGS per unit."""
    price = PRODUCT_MIX_CC * COTTON_CANDY_PRICE + (1 - PRODUCT_MIX_CC) * SHAVED_ICE_PRICE
    cogs  = PRODUCT_MIX_CC * COTTON_CANDY_COGS  + (1 - PRODUCT_MIX_CC) * SHAVED_ICE_COGS
    return price, cogs


def build_monthly_projection(scenario="base"):
    """Build 36-month projection. Returns list of dicts."""
    avg_price, avg_cogs = calc_blended_price()
    monthly_payment = calc_monthly_payment(SBA_LOAN, INTEREST_RATE, LOAN_TERM_MONTHS)

    if scenario == "worst":
        units_base = UNITS_PER_DAY_LOW
    elif scenario == "best":
        units_base = UNITS_PER_DAY_HIGH
    else:
        units_base = UNITS_PER_DAY_BASE

    months = []
    for m in range(1, 37):
        year = (m - 1) // 12 + 1
        month_of_year = ((m - 1) % 12) + 1  # 1-12
        cal_month = month_of_year  # for seasonality

        # Market days per week
        if year == 1:
            if m <= 3:
                market_days_pw = 0  # pre-launch prep
            elif m <= 5:
                market_days_pw = MARKET_DAYS_Y1_RAMP  # swap meets / pop-ups
            else:
                market_days_pw = MARKET_DAYS_Y1_FULL
        elif year == 2:
            market_days_pw = MARKET_DAYS_Y2
        else:
            market_days_pw = MARKET_DAYS_Y3

        # Weeks per month ~4.33
        market_days = market_days_pw * 4.33

        # Seasonal adjustment
        season = SEASONALITY.get(cal_month, 0.8)

        # Experience ramp (you get better over time)
        if year == 1:
            experience = 0.75 + 0.25 * min(1, max(0, (m - 3) / 9))
        elif year == 2:
            experience = 1.0 + 0.05 * (m - 12) / 12  # slight improvement
        else:
            experience = 1.10

        units_per_day = units_base * season * experience
        total_units = round(units_per_day * market_days)

        gross_revenue = total_units * avg_price
        total_cogs = total_units * avg_cogs

        # Private events
        if year == 1:
            events = PRIVATE_EVENTS_Y1[month_of_year - 1]
        elif year == 2:
            events = PRIVATE_EVENTS_Y2[month_of_year - 1]
        else:
            events = PRIVATE_EVENTS_Y3[month_of_year - 1]
        event_revenue = events * PRIVATE_EVENT_REVENUE

        total_revenue = gross_revenue + event_revenue

        # Stall fees (only on market revenue)
        stall_fees = gross_revenue * STALL_FEE_PCT

        # Fixed OPEX
        if m <= 3 and year == 1:
            # Pre-launch: still paying commissary, insurance, some fuel
            fixed_opex = COMMISSARY_RENT + INSURANCE_GL + INSURANCE_AUTO + 100  # minimal
        else:
            fixed_opex = (COMMISSARY_RENT + INSURANCE_GL + INSURANCE_AUTO +
                         FUEL_BASE + GENERATOR_FUEL + MISC_MAINTENANCE +
                         PHONE_PAYMENT_PROC + PACKAGING_OVERHEAD)

        # Labor (Y3 add occasional help)
        labor = 0
        if year == 3 and cal_month in [5,6,7,8,9]:
            labor = 22 * 5 * 4.33  # $22/hr * 5hrs * ~4.33 weeks (one helper on busy days)
        elif year == 3:
            labor = 22 * 5 * 2  # helper 2 days/month off-peak

        total_opex = stall_fees + fixed_opex + labor
        gross_profit = total_revenue - total_cogs
        net_operating = gross_profit - total_opex
        net_after_debt = net_operating - monthly_payment

        months.append({
            "month": m,
            "year": year,
            "cal_month": cal_month,
            "market_days_pw": market_days_pw,
            "market_days": round(market_days, 1),
            "season": season,
            "units_per_day": round(units_per_day),
            "total_units": total_units,
            "gross_revenue": round(gross_revenue),
            "event_revenue": round(event_revenue),
            "total_revenue": round(total_revenue),
            "total_cogs": round(total_cogs),
            "stall_fees": round(stall_fees),
            "fixed_opex": round(fixed_opex),
            "labor": round(labor),
            "total_opex": round(total_opex + total_cogs),
            "gross_profit": round(gross_profit),
            "net_operating": round(net_operating),
            "loan_payment": round(monthly_payment, 2),
            "net_after_debt": round(net_after_debt),
            "events": events,
        })

    return months


def build_amortization(principal, annual_rate, months):
    """Full amortization schedule."""
    r = annual_rate / 12
    payment = calc_monthly_payment(principal, annual_rate, months)
    balance = principal
    schedule = []
    for m in range(1, months + 1):
        interest = balance * r
        principal_pay = payment - interest
        balance -= principal_pay
        if balance < 0:
            balance = 0
        schedule.append({
            "month": m,
            "payment": round(payment, 2),
            "principal": round(principal_pay, 2),
            "interest": round(interest, 2),
            "balance": round(balance, 2),
        })
    return schedule


# ══════════════════════════════════════════════════════════════
#  WORKBOOK BUILDER
# ══════════════════════════════════════════════════════════════

def build_workbook():
    wb = Workbook()

    # ── SHEET 1: ASSUMPTIONS ──────────────────────────────────
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_properties.tabColor = ACCENT_RED
    set_col_widths(ws, {"A": 38, "B": 18, "C": 16, "D": 42})
    ws.sheet_view.showGridLines = False

    write_title(ws, 1, 1, "ARMAGEDDON TREATS — Key Assumptions", 4)
    write_note(ws, 2, 1, "All figures grounded in the Business Plan Audit. Adjust yellow cells to model scenarios.", 4)

    r = 4
    write_section_label(ws, r, 1, "FINANCING", 2)
    r += 1
    assumptions_finance = [
        ("SBA 7(a) Microloan Principal", SBA_LOAN, FMT_USD, ""),
        ("Maximum Interest Rate (Prime + 6.5%)", INTEREST_RATE, FMT_PCT, "Prime 6.75% + 6.5% spread"),
        ("Loan Term", LOAN_TERM_MONTHS, FMT_NUM, "months (10 years)"),
        ("Monthly Debt Service", round(calc_monthly_payment(SBA_LOAN, INTEREST_RATE, LOAN_TERM_MONTHS), 2), FMT_USD_DEC, "Fixed monthly payment"),
    ]
    for label, val, fmt, note in assumptions_finance:
        write_row(ws, r, [label, val, "", note],
                  fmt=[None, fmt, None, None],
                  bold=[True, False, False, False])
        ws.cell(row=r, column=2).fill = fill_pale_yellow
        r += 1

    r += 1
    write_section_label(ws, r, 1, "PRODUCT ECONOMICS", 2)
    r += 1
    headers = ["Metric", "Cotton Candy", "Shaved Ice", "Blended Avg."]
    write_row(ws, r, headers)
    style_header_row(ws, r, 4)
    r += 1
    avg_price, avg_cogs = calc_blended_price()
    product_data = [
        ("Retail Price", COTTON_CANDY_PRICE, SHAVED_ICE_PRICE, avg_price),
        ("COGS per Unit", COTTON_CANDY_COGS, SHAVED_ICE_COGS, avg_cogs),
        ("Gross Margin per Unit",
         COTTON_CANDY_PRICE - COTTON_CANDY_COGS,
         SHAVED_ICE_PRICE - SHAVED_ICE_COGS,
         avg_price - avg_cogs),
        ("Gross Margin %",
         (COTTON_CANDY_PRICE - COTTON_CANDY_COGS) / COTTON_CANDY_PRICE,
         (SHAVED_ICE_PRICE - SHAVED_ICE_COGS) / SHAVED_ICE_PRICE,
         (avg_price - avg_cogs) / avg_price),
    ]
    for row_data in product_data:
        label = row_data[0]
        vals = row_data[1:]
        fmt_type = FMT_PCT if "%" in label else FMT_USD_DEC
        write_row(ws, r, [label] + list(vals),
                  fmt=[None, fmt_type, fmt_type, fmt_type],
                  bold=[True, False, False, True])
        r += 1

    r += 1
    write_section_label(ws, r, 1, "PRODUCT MIX", 2)
    r += 1
    write_row(ws, r, ["Cotton Candy Share", PRODUCT_MIX_CC, "", ""],
              fmt=[None, FMT_PCT, None, None], bold=[True, False, False, False])
    ws.cell(row=r, column=2).fill = fill_pale_yellow
    r += 1
    write_row(ws, r, ["Shaved Ice Share", 1 - PRODUCT_MIX_CC, "", ""],
              fmt=[None, FMT_PCT, None, None], bold=[True, False, False, False])

    r += 2
    write_section_label(ws, r, 1, "OPERATIONAL PARAMETERS", 2)
    r += 1
    headers = ["Parameter", "Year 1", "Year 2", "Year 3"]
    write_row(ws, r, headers)
    style_header_row(ws, r, 4)
    r += 1
    ops_data = [
        ("Market Days / Week (full operation)", MARKET_DAYS_Y1_FULL, MARKET_DAYS_Y2, MARKET_DAYS_Y3),
        ("Base Units / Day (base case)", UNITS_PER_DAY_BASE, UNITS_PER_DAY_BASE, UNITS_PER_DAY_BASE),
        ("Stall Fee (% of gross)", STALL_FEE_PCT, STALL_FEE_PCT, STALL_FEE_PCT),
        ("Private Events / Month (avg.)", "0–2", "1–4", "2–5"),
    ]
    for row_data in ops_data:
        label = row_data[0]
        write_row(ws, r, list(row_data), bold=[True, False, False, False])
        if "%" in label:
            for c in [2, 3, 4]:
                ws.cell(row=r, column=c).number_format = FMT_PCT
        r += 1

    r += 1
    write_section_label(ws, r, 1, "MONTHLY FIXED OPEX", 2)
    r += 1
    opex_items = [
        ("Commissary Rent (parking + wash tier)", COMMISSARY_RENT),
        ("Commercial General Liability Insurance", INSURANCE_GL),
        ("Commercial Auto Insurance", INSURANCE_AUTO),
        ("Fuel (Tow Vehicle)", FUEL_BASE),
        ("Generator Fuel (Propane/Diesel)", GENERATOR_FUEL),
        ("Miscellaneous / Maintenance", MISC_MAINTENANCE),
        ("Phone & Payment Processing", PHONE_PAYMENT_PROC),
        ("Packaging Overhead", PACKAGING_OVERHEAD),
    ]
    total_fixed = 0
    for label, val in opex_items:
        write_row(ws, r, [label, val], fmt=[None, FMT_USD])
        ws.cell(row=r, column=2).fill = fill_pale_yellow
        total_fixed += val
        r += 1
    write_row(ws, r, ["TOTAL MONTHLY FIXED OPEX", total_fixed],
              fmt=[None, FMT_USD], bold=[True, True])
    ws.cell(row=r, column=1).fill = fill_light
    ws.cell(row=r, column=2).fill = fill_light
    ws.cell(row=r, column=2).border = thick_bottom

    r += 2
    write_section_label(ws, r, 1, "SEASONALITY INDEX (Shaved Ice / Cotton Candy)", 4)
    r += 1
    month_names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    write_row(ws, r, ["Month"] + month_names)
    style_header_row(ws, r, 13)
    r += 1
    write_row(ws, r, ["Multiplier"] + [SEASONALITY[i] for i in range(1,13)],
              fmt=[None] + [FMT_NUM_DEC]*12)

    # ── SHEET 2: STARTUP COSTS ────────────────────────────────
    ws2 = wb.create_sheet("Startup Costs")
    ws2.sheet_properties.tabColor = "2D2D2D"
    set_col_widths(ws2, {"A": 40, "B": 16, "C": 14, "D": 42})
    ws2.sheet_view.showGridLines = False

    write_title(ws2, 1, 1, "Startup Capital Allocation — $14,000 SBA 7(a)", 4)
    write_note(ws2, 2, 1, "Zero margin for error. Every dollar mapped to a mandatory function.", 4)

    r = 4
    headers = ["Expense Category", "Allocation", "% of Total", "Strategic Justification"]
    write_row(ws2, r, headers)
    style_header_row(ws2, r, 4)
    r += 1

    justifications = {
        "Concession Trailer (10ft import)": "Compliant imported trailer with basic electrical and multi-compartment sinks",
        "Equipment & Branding (wrap, machines)": "ANSI/NSF certified ice shaver, cotton candy spinner, matte-black vinyl wrap",
        "Permits & Health Licensing": "LACDPH plan check ($347), MFF annual fee ($598), SM Vendor License (~$175)",
        "Initial Insurance Premiums": "Down payments on Commercial GL and Commercial Auto policies",
        "Initial Inventory & Packaging": "Bulk organic floss sugar, pure cane syrups, biodegradable PLA flower cups",
        "Commissary Deposit / First Month": "Secure LACDPH Verification of Proper Food Vehicle Storage form",
        "Working Capital / Contingency": "Cash reserve for unforeseen compliance modifications, fuel, or delays",
    }

    for item, cost in STARTUP_COSTS.items():
        pct = cost / SBA_LOAN
        write_row(ws2, r, [item, cost, pct, justifications.get(item, "")],
                  fmt=[None, FMT_USD, FMT_PCT, None])
        r += 1

    # Total row
    write_row(ws2, r, ["TOTAL CAPITAL DEPLOYMENT", SBA_LOAN, 1.0, "Fully exhausted SBA 7(a) Microloan"],
              fmt=[None, FMT_USD, FMT_PCT, None], bold=True)
    style_total_row(ws2, r, 4)

    # ── SHEET 3: MONTHLY PROJECTIONS (36 months) ──────────────
    ws3 = wb.create_sheet("Monthly Projections")
    ws3.sheet_properties.tabColor = MEDIUM_BLUE
    ws3.sheet_view.showGridLines = False

    headers_monthly = [
        "Month #", "Year", "Calendar Mo.", "Mkt Days/Wk",
        "Total Mkt Days", "Seasonal Idx", "Units/Day", "Total Units",
        "Market Revenue", "Event Revenue", "TOTAL REVENUE",
        "COGS", "Stall Fees", "Fixed OPEX", "Labor",
        "TOTAL EXPENSES", "Gross Profit", "Net Operating",
        "Loan Payment", "NET AFTER DEBT"
    ]

    write_title(ws3, 1, 1, "36-Month Financial Projections — Base Case", 10)
    write_note(ws3, 2, 1, f"Base case: {UNITS_PER_DAY_BASE} units/day avg. | Blended price ${calc_blended_price()[0]:.2f} | Seasonal adjustment applied", 10)

    r = 4
    write_row(ws3, r, headers_monthly)
    style_header_row(ws3, r, len(headers_monthly))

    col_widths_monthly = {
        "A": 9, "B": 7, "C": 13, "D": 12, "E": 12, "F": 11,
        "G": 10, "H": 10, "I": 14, "J": 13, "K": 14,
        "L": 10, "M": 11, "N": 11, "O": 10,
        "P": 14, "Q": 13, "R": 13, "S": 12, "T": 14
    }
    set_col_widths(ws3, col_widths_monthly)

    base_months = build_monthly_projection("base")
    month_names_full = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    r = 5
    for md in base_months:
        cal_name = month_names_full[md["cal_month"] - 1]
        year_label = md["year"]
        row_data = [
            md["month"], year_label, cal_name, md["market_days_pw"],
            md["market_days"], md["season"], md["units_per_day"], md["total_units"],
            md["gross_revenue"], md["event_revenue"], md["total_revenue"],
            md["total_cogs"], md["stall_fees"], md["fixed_opex"], md["labor"],
            md["total_opex"], md["gross_profit"], md["net_operating"],
            md["loan_payment"], md["net_after_debt"]
        ]
        fmts = [
            FMT_NUM, FMT_NUM, None, FMT_NUM_DEC,
            FMT_NUM_DEC, FMT_NUM_DEC, FMT_NUM, FMT_NUM,
            FMT_USD, FMT_USD, FMT_USD,
            FMT_USD, FMT_USD, FMT_USD, FMT_USD,
            FMT_USD, FMT_USD, FMT_USD,
            FMT_USD, FMT_USD
        ]
        write_row(ws3, r, row_data, fmt=fmts)

        # Color net after debt
        net_cell = ws3.cell(row=r, column=20)
        if md["net_after_debt"] < 0:
            net_cell.font = font_negative
            net_cell.fill = fill_pale_red
        else:
            net_cell.font = font_positive
            net_cell.fill = fill_pale_green

        # Alternate year shading
        if md["year"] == 2:
            for c in range(1, 9):
                ws3.cell(row=r, column=c).fill = fill_warm

        r += 1

    # Year subtotals
    for yr in [1, 2, 3]:
        yr_data = [m for m in base_months if m["year"] == yr]
        write_row(ws3, r, [
            None, f"YEAR {yr}", "TOTAL", None,
            sum(m["market_days"] for m in yr_data), None,
            None, sum(m["total_units"] for m in yr_data),
            sum(m["gross_revenue"] for m in yr_data),
            sum(m["event_revenue"] for m in yr_data),
            sum(m["total_revenue"] for m in yr_data),
            sum(m["total_cogs"] for m in yr_data),
            sum(m["stall_fees"] for m in yr_data),
            sum(m["fixed_opex"] for m in yr_data),
            sum(m["labor"] for m in yr_data),
            sum(m["total_opex"] for m in yr_data),
            sum(m["gross_profit"] for m in yr_data),
            sum(m["net_operating"] for m in yr_data),
            sum(m["loan_payment"] for m in yr_data),
            sum(m["net_after_debt"] for m in yr_data),
        ], fmt=[
            None, None, None, None,
            FMT_NUM_DEC, None, None, FMT_NUM,
            FMT_USD, FMT_USD, FMT_USD,
            FMT_USD, FMT_USD, FMT_USD, FMT_USD,
            FMT_USD, FMT_USD, FMT_USD,
            FMT_USD, FMT_USD
        ], bold=True)
        style_total_row(ws3, r, 20, fill=fill_subheader if yr < 3 else fill_accent)
        r += 1

    # Grand total
    all_data = base_months
    write_row(ws3, r, [
        None, "3-YEAR", "GRAND TOTAL", None, None, None, None,
        sum(m["total_units"] for m in all_data),
        sum(m["gross_revenue"] for m in all_data),
        sum(m["event_revenue"] for m in all_data),
        sum(m["total_revenue"] for m in all_data),
        sum(m["total_cogs"] for m in all_data),
        sum(m["stall_fees"] for m in all_data),
        sum(m["fixed_opex"] for m in all_data),
        sum(m["labor"] for m in all_data),
        sum(m["total_opex"] for m in all_data),
        sum(m["gross_profit"] for m in all_data),
        sum(m["net_operating"] for m in all_data),
        sum(m["loan_payment"] for m in all_data),
        sum(m["net_after_debt"] for m in all_data),
    ], fmt=[
        None, None, None, None, None, None, None, FMT_NUM,
        FMT_USD, FMT_USD, FMT_USD,
        FMT_USD, FMT_USD, FMT_USD, FMT_USD,
        FMT_USD, FMT_USD, FMT_USD,
        FMT_USD, FMT_USD
    ], bold=True)
    style_total_row(ws3, r, 20, fill=fill_accent)

    # Freeze panes
    ws3.freeze_panes = "A5"

    # ── SHEET 4: ANNUAL SUMMARY ───────────────────────────────
    ws4 = wb.create_sheet("Annual Summary")
    ws4.sheet_properties.tabColor = DARK_GREEN
    ws4.sheet_view.showGridLines = False
    set_col_widths(ws4, {"A": 32, "B": 16, "C": 16, "D": 16, "E": 16})

    write_title(ws4, 1, 1, "3-Year Annual P&L Summary", 5)

    r = 3
    headers = ["", "Year 1", "Year 2", "Year 3", "3-Year Total"]
    write_row(ws4, r, headers)
    style_header_row(ws4, r, 5)

    # Compute annual totals
    annual = {}
    for yr in [1, 2, 3]:
        yr_data = [m for m in base_months if m["year"] == yr]
        annual[yr] = {
            "total_revenue": sum(m["total_revenue"] for m in yr_data),
            "gross_revenue": sum(m["gross_revenue"] for m in yr_data),
            "event_revenue": sum(m["event_revenue"] for m in yr_data),
            "total_cogs": sum(m["total_cogs"] for m in yr_data),
            "gross_profit": sum(m["gross_profit"] for m in yr_data),
            "stall_fees": sum(m["stall_fees"] for m in yr_data),
            "fixed_opex": sum(m["fixed_opex"] for m in yr_data),
            "labor": sum(m["labor"] for m in yr_data),
            "total_opex": sum(m["total_opex"] for m in yr_data),
            "net_operating": sum(m["net_operating"] for m in yr_data),
            "loan_payments": sum(m["loan_payment"] for m in yr_data),
            "net_after_debt": sum(m["net_after_debt"] for m in yr_data),
            "total_units": sum(m["total_units"] for m in yr_data),
            "market_days": sum(m["market_days"] for m in yr_data),
        }

    t = {k: sum(annual[yr][k] for yr in [1,2,3]) for k in annual[1].keys()}

    r = 4
    pnl_rows = [
        ("REVENUE", None, True, fill_light),
        ("  Market Revenue", "gross_revenue", False, None),
        ("  Private Event Revenue", "event_revenue", False, None),
        ("Total Revenue", "total_revenue", True, fill_pale_blue),
        ("", None, False, None),
        ("COST OF GOODS SOLD", None, True, fill_light),
        ("  Product COGS", "total_cogs", False, None),
        ("Gross Profit", "gross_profit", True, fill_pale_green),
        ("", None, False, None),
        ("OPERATING EXPENSES", None, True, fill_light),
        ("  Stall Fees (12.5% of market rev.)", "stall_fees", False, None),
        ("  Fixed OPEX", "fixed_opex", False, None),
        ("  Labor", "labor", False, None),
        ("Total Operating Expenses", "total_opex", True, fill_pale_red),
        ("", None, False, None),
        ("Net Operating Income", "net_operating", True, fill_pale_green),
        ("  Less: SBA Loan Payments", "loan_payments", False, None),
        ("NET INCOME (After Debt Service)", "net_after_debt", True, None),
    ]

    for label, key, bold, fill in pnl_rows:
        if not label:
            r += 1
            continue
        row_vals = [label]
        for yr in [1, 2, 3]:
            row_vals.append(annual[yr].get(key, None) if key else None)
        row_vals.append(t.get(key, None) if key else None)

        write_row(ws4, r, row_vals,
                  fmt=[None, FMT_USD, FMT_USD, FMT_USD, FMT_USD],
                  bold=[bold, bold, bold, bold, bold])
        if fill:
            for c in range(1, 6):
                ws4.cell(row=r, column=c).fill = fill
        if key == "net_after_debt":
            style_total_row(ws4, r, 5)
        r += 1

    # KPI section
    r += 2
    write_section_label(ws4, r, 1, "KEY PERFORMANCE INDICATORS", 5)
    r += 1
    write_row(ws4, r, ["", "Year 1", "Year 2", "Year 3", "3-Year"])
    style_header_row(ws4, r, 5)
    r += 1

    kpis = [
        ("Total Units Sold", "total_units", FMT_NUM),
        ("Total Market Days", "market_days", FMT_NUM),
        ("Avg. Revenue / Market Day", None, FMT_USD),
        ("Gross Margin %", None, FMT_PCT),
        ("Operating Margin %", None, FMT_PCT),
    ]
    for label, key, fmt in kpis:
        row_vals = [label]
        for yr in [1, 2, 3]:
            a = annual[yr]
            if key:
                row_vals.append(a[key])
            elif "Revenue / Market" in label:
                row_vals.append(round(a["total_revenue"] / max(a["market_days"], 1)))
            elif "Gross Margin" in label:
                row_vals.append(a["gross_profit"] / max(a["total_revenue"], 1))
            elif "Operating" in label:
                row_vals.append(a["net_operating"] / max(a["total_revenue"], 1))
        # Total column
        if key:
            row_vals.append(t[key])
        elif "Revenue / Market" in label:
            row_vals.append(round(t["total_revenue"] / max(t["market_days"], 1)))
        elif "Gross Margin" in label:
            row_vals.append(t["gross_profit"] / max(t["total_revenue"], 1))
        elif "Operating" in label:
            row_vals.append(t["net_operating"] / max(t["total_revenue"], 1))

        write_row(ws4, r, row_vals, fmt=[None, fmt, fmt, fmt, fmt], bold=[True]+[False]*4)
        r += 1

    # ── SHEET 5: SCENARIO ANALYSIS ────────────────────────────
    ws5 = wb.create_sheet("Scenario Analysis")
    ws5.sheet_properties.tabColor = "E67E22"
    ws5.sheet_view.showGridLines = False
    set_col_widths(ws5, {"A": 32, "B": 16, "C": 16, "D": 16})

    write_title(ws5, 1, 1, "Scenario Comparison: Worst / Base / Best", 4)
    write_note(ws5, 2, 1, f"Worst: {UNITS_PER_DAY_LOW} units/day | Base: {UNITS_PER_DAY_BASE} units/day | Best: {UNITS_PER_DAY_HIGH} units/day", 4)

    scenarios = {}
    for sc in ["worst", "base", "best"]:
        months_data = build_monthly_projection(sc)
        sc_annual = {}
        for yr in [1, 2, 3]:
            yr_data = [m for m in months_data if m["year"] == yr]
            sc_annual[yr] = {
                "total_revenue": sum(m["total_revenue"] for m in yr_data),
                "total_cogs": sum(m["total_cogs"] for m in yr_data),
                "gross_profit": sum(m["gross_profit"] for m in yr_data),
                "total_opex": sum(m["total_opex"] for m in yr_data),
                "net_operating": sum(m["net_operating"] for m in yr_data),
                "loan_payments": sum(m["loan_payment"] for m in yr_data),
                "net_after_debt": sum(m["net_after_debt"] for m in yr_data),
                "total_units": sum(m["total_units"] for m in yr_data),
            }
        sc_annual["total"] = {k: sum(sc_annual[yr][k] for yr in [1,2,3]) for k in sc_annual[1].keys()}
        scenarios[sc] = sc_annual

    for yr_label, yr_key in [("Year 1", 1), ("Year 2", 2), ("Year 3", 3), ("3-Year Total", "total")]:
        r_start = 4 if yr_key == 1 else r + 1
        r = r_start

        write_section_label(ws5, r, 1, yr_label, 4)
        r += 1
        write_row(ws5, r, ["", "Worst Case", "Base Case", "Best Case"])
        style_header_row(ws5, r, 4)
        r += 1

        fills = [None, fill_pale_red, fill_pale_yellow, fill_pale_green]
        for metric, key, fmt in [
            ("Total Units Sold", "total_units", FMT_NUM),
            ("Total Revenue", "total_revenue", FMT_USD),
            ("COGS", "total_cogs", FMT_USD),
            ("Total Expenses (COGS + OPEX)", "total_opex", FMT_USD),
            ("Net Operating Income", "net_operating", FMT_USD),
            ("Net After Debt Service", "net_after_debt", FMT_USD),
        ]:
            row_vals = [metric]
            for sc in ["worst", "base", "best"]:
                row_vals.append(scenarios[sc][yr_key][key])
            write_row(ws5, r, row_vals, fmt=[None, fmt, fmt, fmt], bold=[True]+[False]*3,
                      fill=fills)
            r += 1

    # ── SHEET 6: LOAN AMORTIZATION ────────────────────────────
    ws6 = wb.create_sheet("Loan Amortization")
    ws6.sheet_properties.tabColor = "8E44AD"
    ws6.sheet_view.showGridLines = False
    set_col_widths(ws6, {"A": 10, "B": 8, "C": 14, "D": 14, "E": 14, "F": 16})

    write_title(ws6, 1, 1, "SBA 7(a) Loan Amortization Schedule", 6)
    payment = calc_monthly_payment(SBA_LOAN, INTEREST_RATE, LOAN_TERM_MONTHS)
    total_paid = payment * LOAN_TERM_MONTHS
    total_interest = total_paid - SBA_LOAN
    write_note(ws6, 2, 1,
        f"Principal: $14,000 | Rate: 13.25% | Term: 120 months | "
        f"Payment: ${payment:.2f}/mo | Total Interest: ${total_interest:,.0f} | "
        f"Total Paid: ${total_paid:,.0f}", 6)

    r = 4
    headers = ["Month", "Year", "Payment", "Principal", "Interest", "Balance"]
    write_row(ws6, r, headers)
    style_header_row(ws6, r, 6)

    amort = build_amortization(SBA_LOAN, INTEREST_RATE, LOAN_TERM_MONTHS)
    r = 5
    for entry in amort:
        yr = math.ceil(entry["month"] / 12)
        write_row(ws6, r, [
            entry["month"], yr, entry["payment"],
            entry["principal"], entry["interest"], entry["balance"]
        ], fmt=[FMT_NUM, FMT_NUM, FMT_USD_DEC, FMT_USD_DEC, FMT_USD_DEC, FMT_USD_DEC])

        if entry["month"] % 12 == 0:
            for c in range(1, 7):
                ws6.cell(row=r, column=c).border = thick_bottom
        r += 1

    # Year-end summaries at bottom
    r += 1
    write_section_label(ws6, r, 1, "ANNUAL SUMMARY", 6)
    r += 1
    write_row(ws6, r, ["Year", "", "Total Paid", "Principal", "Interest", "End Balance"])
    style_header_row(ws6, r, 6)
    r += 1

    for yr in range(1, 11):
        yr_entries = [e for e in amort if math.ceil(e["month"] / 12) == yr]
        write_row(ws6, r, [
            yr, None,
            sum(e["payment"] for e in yr_entries),
            sum(e["principal"] for e in yr_entries),
            sum(e["interest"] for e in yr_entries),
            yr_entries[-1]["balance"]
        ], fmt=[FMT_NUM, None, FMT_USD_DEC, FMT_USD_DEC, FMT_USD_DEC, FMT_USD_DEC])
        r += 1

    ws6.freeze_panes = "A5"

    # ── SHEET 7: BREAK-EVEN ANALYSIS ─────────────────────────
    ws7 = wb.create_sheet("Break-Even")
    ws7.sheet_properties.tabColor = "27AE60"
    ws7.sheet_view.showGridLines = False
    set_col_widths(ws7, {"A": 38, "B": 18, "C": 18, "D": 36})

    write_title(ws7, 1, 1, "Break-Even Analysis", 4)

    avg_price, avg_cogs = calc_blended_price()
    monthly_fixed = (COMMISSARY_RENT + INSURANCE_GL + INSURANCE_AUTO +
                     FUEL_BASE + GENERATOR_FUEL + MISC_MAINTENANCE +
                     PHONE_PAYMENT_PROC + PACKAGING_OVERHEAD)
    monthly_debt = calc_monthly_payment(SBA_LOAN, INTEREST_RATE, LOAN_TERM_MONTHS)

    # Contribution margin per unit after stall fee
    net_price = avg_price * (1 - STALL_FEE_PCT)
    contribution = net_price - avg_cogs

    be_units_monthly = math.ceil((monthly_fixed + monthly_debt) / contribution)
    be_revenue_monthly = be_units_monthly * avg_price
    be_units_daily_2days = math.ceil(be_units_monthly / (2 * 4.33))
    be_units_daily_3days = math.ceil(be_units_monthly / (3 * 4.33))

    r = 3
    write_section_label(ws7, r, 1, "UNIT CONTRIBUTION ANALYSIS", 4)
    r += 1

    be_data = [
        ("Blended Average Price", avg_price, FMT_USD_DEC, ""),
        ("Less: Stall Fee (12.5%)", avg_price * STALL_FEE_PCT, FMT_USD_DEC, "Variable % of gross revenue"),
        ("Net Revenue per Unit", net_price, FMT_USD_DEC, ""),
        ("Less: COGS per Unit", avg_cogs, FMT_USD_DEC, ""),
        ("CONTRIBUTION MARGIN / UNIT", contribution, FMT_USD_DEC, "What each unit contributes to fixed costs"),
        ("Contribution Margin %", contribution / avg_price, FMT_PCT, "As % of retail price"),
    ]
    for label, val, fmt, note in be_data:
        write_row(ws7, r, [label, val, "", note], fmt=[None, fmt, None, None],
                  bold=[True if "CONTRIBUTION" in label or "Net Revenue" in label else False,
                        True if "CONTRIBUTION" in label else False, False, False])
        r += 1

    r += 1
    write_section_label(ws7, r, 1, "MONTHLY BREAK-EVEN TARGETS", 4)
    r += 1
    write_row(ws7, r, ["Total Monthly Fixed Costs", monthly_fixed, "", "Commissary + insurance + fuel + misc"])
    style_data_cell(ws7.cell(row=r, column=2), FMT_USD, bold=True)
    r += 1
    write_row(ws7, r, ["Monthly Loan Payment", round(monthly_debt, 2), "", "SBA 7(a) debt service"])
    style_data_cell(ws7.cell(row=r, column=2), FMT_USD_DEC)
    r += 1
    write_row(ws7, r, ["Total Monthly Nut (Fixed + Debt)", round(monthly_fixed + monthly_debt), "", "Must cover before profit"])
    style_data_cell(ws7.cell(row=r, column=2), FMT_USD, bold=True)
    ws7.cell(row=r, column=1).fill = fill_pale_red
    ws7.cell(row=r, column=2).fill = fill_pale_red
    r += 2

    write_row(ws7, r, ["Break-Even Units / Month", be_units_monthly, "", "At current blended price & margin"])
    style_data_cell(ws7.cell(row=r, column=2), FMT_NUM, bold=True)
    ws7.cell(row=r, column=1).fill = fill_pale_yellow
    ws7.cell(row=r, column=2).fill = fill_pale_yellow
    r += 1
    write_row(ws7, r, ["Break-Even Revenue / Month", round(be_revenue_monthly), "", ""])
    style_data_cell(ws7.cell(row=r, column=2), FMT_USD, bold=True)
    r += 1
    write_row(ws7, r, ["Break-Even Units / Day (2 days/wk)", be_units_daily_2days, "", "Year 1 target"])
    style_data_cell(ws7.cell(row=r, column=2), FMT_NUM, bold=True)
    ws7.cell(row=r, column=1).fill = fill_pale_green
    ws7.cell(row=r, column=2).fill = fill_pale_green
    r += 1
    write_row(ws7, r, ["Break-Even Units / Day (3 days/wk)", be_units_daily_3days, "", "Year 3 target"])
    style_data_cell(ws7.cell(row=r, column=2), FMT_NUM, bold=True)
    ws7.cell(row=r, column=1).fill = fill_pale_green
    ws7.cell(row=r, column=2).fill = fill_pale_green

    r += 2
    write_section_label(ws7, r, 1, "STARTUP COST RECOVERY", 4)
    r += 1
    # Find month where cumulative net > 0
    cumulative = 0
    recovery_month = None
    for md in base_months:
        cumulative += md["net_after_debt"]
        if cumulative > 0 and recovery_month is None:
            recovery_month = md["month"]

    write_row(ws7, r, ["Initial Investment (SBA Loan)", SBA_LOAN, "", ""])
    style_data_cell(ws7.cell(row=r, column=2), FMT_USD)
    r += 1
    if recovery_month:
        write_row(ws7, r, ["Month of Full Cost Recovery (Base Case)", recovery_month, "",
                          f"Cumulative net profit turns positive in month {recovery_month}"])
    else:
        write_row(ws7, r, ["Month of Full Cost Recovery (Base Case)", "Beyond 36 mo.", "", ""])
    style_data_cell(ws7.cell(row=r, column=2), FMT_NUM, bold=True)
    ws7.cell(row=r, column=1).fill = fill_pale_blue
    ws7.cell(row=r, column=2).fill = fill_pale_blue

    r += 1
    cumulative = 0
    for md in base_months:
        cumulative += md["net_after_debt"]
    write_row(ws7, r, ["Cumulative Net Profit (36 months, base)", round(cumulative), "", ""])
    style_data_cell(ws7.cell(row=r, column=2), FMT_USD, bold=True)

    # ── SHEET 8: SENSITIVITY TABLE ────────────────────────────
    ws8 = wb.create_sheet("Sensitivity")
    ws8.sheet_properties.tabColor = "D35400"
    ws8.sheet_view.showGridLines = False
    set_col_widths(ws8, {"A": 20, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14})

    write_title(ws8, 1, 1, "Sensitivity Analysis: Units/Day vs. Market Days/Week", 7)
    write_note(ws8, 2, 1, "Shows Year 2 NET OPERATING INCOME at various volume/frequency combinations (annual, base seasonality applied)", 7)

    # Build a grid: rows = units/day, cols = market days/week
    units_range = [30, 40, 50, 60, 70, 80, 100, 120, 150]
    days_range = [1, 1.5, 2, 2.5, 3, 3.5]

    r = 4
    write_row(ws8, r, ["Units/Day \\ Days/Wk"] + [f"{d}" for d in days_range])
    style_header_row(ws8, r, 1 + len(days_range))

    r = 5
    for units in units_range:
        row_vals = [units]
        for days_pw in days_range:
            # Simulate Year 2 annual income
            annual_income = 0
            for cal_month in range(1, 13):
                season = SEASONALITY[cal_month]
                market_days = days_pw * 4.33
                day_units = round(units * season)
                total_units = round(day_units * market_days)
                gross_rev = total_units * avg_price
                cogs = total_units * avg_cogs
                stall = gross_rev * STALL_FEE_PCT
                fixed = monthly_fixed
                net = gross_rev - cogs - stall - fixed
                annual_income += net
            row_vals.append(round(annual_income))

        write_row(ws8, r, row_vals,
                  fmt=[FMT_NUM] + [FMT_USD] * len(days_range))

        # Color code
        for c_idx in range(1, len(days_range) + 1):
            cell = ws8.cell(row=r, column=c_idx + 1)
            val = cell.value
            if val is not None:
                if val < 0:
                    cell.fill = fill_pale_red
                    cell.font = font_negative
                elif val < 5000:
                    cell.fill = fill_pale_yellow
                else:
                    cell.fill = fill_pale_green
                    cell.font = font_positive

        # Highlight the base case row
        if units == UNITS_PER_DAY_BASE:
            ws8.cell(row=r, column=1).fill = fill_pale_blue
            ws8.cell(row=r, column=1).font = font_body_bold

        r += 1

    r += 1
    write_note(ws8, r, 1, "Green = profitable | Yellow = marginal ($0\u2013$5k) | Red = loss. Does NOT include loan payments or labor.", 7)

    # ── SHEET 9: CASH FLOW RUNWAY ─────────────────────────────
    ws9 = wb.create_sheet("Cash Flow Runway")
    ws9.sheet_properties.tabColor = "1ABC9C"
    ws9.sheet_view.showGridLines = False
    set_col_widths(ws9, {"A": 10, "B": 10, "C": 14, "D": 14, "E": 14, "F": 16, "G": 14})

    write_title(ws9, 1, 1, "Cumulative Cash Flow Runway — Base Case", 7)
    write_note(ws9, 2, 1, "Starting cash = $0 after $14k loan fully deployed to startup costs. Tracks monthly cash position.", 7)

    r = 4
    headers = ["Month", "Year", "Revenue", "Total Costs", "Net Cash Flow", "Cumulative Cash", "Status"]
    write_row(ws9, r, headers)
    style_header_row(ws9, r, 7)

    r = 5
    cumulative_cash = 0
    lowest_cash = 0
    for md in base_months:
        net_cf = md["net_after_debt"]
        cumulative_cash += net_cf
        if cumulative_cash < lowest_cash:
            lowest_cash = cumulative_cash

        status = "Positive" if cumulative_cash >= 0 else "Deficit"
        month_label = month_names_full[md["cal_month"] - 1]

        write_row(ws9, r, [
            md["month"], md["year"], md["total_revenue"],
            md["total_opex"] + md["loan_payment"],
            net_cf, round(cumulative_cash), status
        ], fmt=[FMT_NUM, None, FMT_USD, FMT_USD, FMT_USD, FMT_USD, None])

        # Color cumulative
        cum_cell = ws9.cell(row=r, column=6)
        if cumulative_cash < 0:
            cum_cell.fill = fill_pale_red
            cum_cell.font = font_negative
        else:
            cum_cell.fill = fill_pale_green
            cum_cell.font = font_positive

        status_cell = ws9.cell(row=r, column=7)
        if status == "Deficit":
            status_cell.fill = fill_pale_red
            status_cell.font = font_negative
        else:
            status_cell.fill = fill_pale_green
            status_cell.font = font_positive

        r += 1

    r += 1
    write_section_label(ws9, r, 1, "CASH FLOW SUMMARY", 7)
    r += 1
    write_row(ws9, r, ["Lowest Cash Position", None, None, None, None, round(lowest_cash), None],
              fmt=[None, None, None, None, None, FMT_USD, None], bold=True)
    ws9.cell(row=r, column=6).fill = fill_pale_red
    r += 1
    write_row(ws9, r, ["Final Cash Position (Mo. 36)", None, None, None, None, round(cumulative_cash), None],
              fmt=[None, None, None, None, None, FMT_USD, None], bold=True)
    ws9.cell(row=r, column=6).fill = fill_pale_green if cumulative_cash > 0 else fill_pale_red

    ws9.freeze_panes = "A5"

    # ── SAVE ──────────────────────────────────────────────────
    filename = "Armageddon_Treats_Financial_Projections.xlsx"
    wb.save(filename)
    print(f"Workbook saved: {filename}")
    print(f"\nSheets: {[s.title for s in wb.worksheets]}")

    # Print summary
    print("\n-- QUICK SUMMARY --")
    for yr in [1, 2, 3]:
        a = annual[yr]
        print(f"  Year {yr}: Revenue ${a['total_revenue']:,} | Net Operating ${a['net_operating']:,} | After Debt ${a['net_after_debt']:,}")
    print(f"  3-Year Total: Revenue ${t['total_revenue']:,} | Net Operating ${t['net_operating']:,} | After Debt ${t['net_after_debt']:,}")
    print(f"  Monthly Loan Payment: ${monthly_debt:.2f}")
    print(f"  Break-Even: {be_units_monthly} units/month ({be_units_daily_2days}/day @ 2 days/wk)")
    if recovery_month:
        print(f"  Cost Recovery: Month {recovery_month}")


if __name__ == "__main__":
    build_workbook()
